"""
Import Business Development Executive (BDE) candidate exports (Naukri + Apna style).

Usage (from BE/):
  python -m app.import_bde
  python -m app.import_bde --reset
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

import openpyxl
from sqlalchemy import select, func

from app.database import async_session, engine, Base
from app.models.hiring import Candidate, HiringRole

ROLE_ID = "business_development_executive"
ROLE_NAME = "Business Development Executive"

DOWNLOADS = Path.home() / "Downloads"
DEFAULT_FILES = [
    DOWNLOADS / "Application_10910516_160.xlsx",
    DOWNLOADS / "BDE_759813_10-18-2021_Hyderabad.xlsx",
    DOWNLOADS / "BDE_402355936_12-16-2021.xlsx",
    DOWNLOADS / "BDE's.xlsx",
    DOWNLOADS / "Business_Development_Executive.xlsx",
    DOWNLOADS / "Business-Development_20230901113534_155.xlsx",
    DOWNLOADS / "Business-Development_20230909140133_160.xlsx",
    DOWNLOADS / "Business-Development_20240428223327_40.xlsx",
    DOWNLOADS / "Business-Development_20240428224527_1.xlsx",
]


def _get(row: dict, *keys: str):
    lower = {str(k).lstrip("\ufeff").strip().lower(): v for k, v in row.items() if k is not None}
    for key in keys:
        if key.lower() in lower:
            v = lower[key.lower()]
            if v is None:
                continue
            s = str(v).strip()
            if s.endswith(".0") and re.fullmatch(r"\d+\.0", s):
                s = s[:-2]
            if s and s.lower() not in {"none", "nan", "null", "na", "n/a"}:
                return s
    return None


def _phone(raw: str | None) -> str | None:
    if not raw:
        return None
    # take first number if multiple
    parts = re.split(r"[,;/]", raw)
    for p in parts:
        d = re.sub(r"\D", "", p)
        if len(d) >= 10:
            return d[-10:]
    d = re.sub(r"\D", "", raw)
    return d[-10:] if len(d) >= 10 else (d or None)


def _email(raw: str | None) -> str | None:
    if not raw:
        return None
    parts = re.split(r"[,;\s]+", raw)
    for p in parts:
        if "@" in p:
            return p.strip().lower()
    return raw.strip().lower() or None


def _has_exp(exp: str | None) -> str | None:
    if not exp:
        return None
    if re.search(r"(?i)fresher|0\s*year", exp) and not re.search(r"(?i)[1-9]", exp):
        return "No"
    if re.search(r"(?i)\d|<1", exp):
        return "Yes"
    return None


def _status(*parts: str | None) -> str:
    blob = " ".join(p for p in parts if p).lower()
    if "reject" in blob or "not interested" in blob:
        return "rejected"
    if "shortlist" in blob or "selected" in blob:
        return "shortlisted"
    if "interview" in blob:
        return "interview"
    if "hold" in blob or "pending" in blob or "callback" in blob or "call back" in blob or "not answering" in blob:
        return "on_hold"
    return "new"


def load_xlsx_all_sheets(path: Path) -> list[tuple[str, list[dict]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple[str, list[dict]]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
        # skip empty header sheets
        if not any(headers):
            continue
        rows = []
        for raw in it:
            if not raw or all(c is None or str(c).strip() == "" for c in raw):
                continue
            rows.append({headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))})
        if rows:
            out.append((sn, rows))
    wb.close()
    return out


def is_naukri_headers(row: dict) -> bool:
    keys = {str(k).lower() for k in row.keys()}
    return "email id" in keys or "job title" in keys or "current location" in keys


def is_apna_headers(row: dict) -> bool:
    keys = {str(k).lower() for k in row.keys()}
    return "phone number" in keys and "name" in keys and ("area" in keys or "education" in keys)


def row_to_payload(row: dict, source: str) -> dict | None:
    name = _get(row, "Name")
    email = _email(_get(row, "Email ID", "Email"))
    phone = _phone(_get(row, "Phone Number", "Phone"))
    if not name and not email and not phone:
        return None

    if is_naukri_headers(row) or email:
        city = _get(row, "Current Location", "City")
        exp = _get(row, "Total Experience", "Experience")
        company = _get(row, "Curr. Company name", "Company")
        designation = _get(row, "Curr. Company Designation", "Current Job", "Role")
        skills = _get(row, "Key Skills", "Skills")
        summary = _get(row, "Summary")
        headline = _get(row, "Resume Headline")
        ug = _get(row, "Under Graduation degree")
        ug_inst = _get(row, "UG University/institute Name")
        ug_year = _get(row, "UG Graduation year")
        availability = _get(row, "Notice period/ Availability to join", "Notice Period")
        applied = _get(row, "Date of application", "Date", "Applied On")
        notes_extra = _get(row, "Notes")
        feedback = _get(row, "Feedback")
        status_raw = _get(row, "Status")
        job_title = _get(row, "Job Title")
        industry = _get(row, "Industry")
        department = _get(row, "Department", "Functional Area")

        notes = [f"Source: {source}"]
        if job_title:
            notes.append(f"Job: {job_title}")
        if notes_extra:
            notes.append(f"Notes: {notes_extra}")
        if feedback:
            notes.append(f"Feedback: {feedback}")
        if status_raw:
            notes.append(f"Export status: {status_raw}")
        if industry:
            notes.append(f"Industry: {industry}")
        if department:
            notes.append(f"Department: {department}")

        key = (email or f"{phone or ''}|{(name or '').lower()}")
        cid = f"bde_{hashlib.sha1(key.encode()).hexdigest()[:12]}"

        return {
            "id": cid,
            "name": name,
            "email": email,
            "phone": phone,
            "city": city.split(",")[0].strip() if city else None,
            "experience_duration": exp,
            "has_work_experience": _has_exp(exp),
            "companies": company,
            "latest_company": company,
            "job_titles": designation,
            "latest_role": designation,
            "other_skills": skills,
            "career_objective": headline or summary,
            "work_experience_detail": summary,
            "degree": ug,
            "institute": ug_inst,
            "graduation_year": ug_year,
            "availability": availability,
            "applied_at": applied,
            "status": _status(status_raw, notes_extra, feedback),
            "notes": "\n".join(notes),
            "tags": ["bde", "business_development", "excel_import"],
        }

    # Apna-style
    city = _get(row, "Area", "Current Location")
    company = _get(row, "Company")
    job = _get(row, "Current Job")
    education = _get(row, "Education")
    exp = _get(row, "Experience", "Total Experience")
    status_raw = _get(row, "Status")
    feedback = _get(row, "Feedback")
    applied = _get(row, "Applied On", "Date of application", "Date")

    key = f"{phone or ''}|{(name or '').lower()}"
    cid = f"bde_{hashlib.sha1(key.encode()).hexdigest()[:12]}"
    notes = [f"Source: {source}"]
    if feedback:
        notes.append(f"Feedback: {feedback}")
    if status_raw:
        notes.append(f"Apna/status: {status_raw}")

    return {
        "id": cid,
        "name": name,
        "email": email,
        "phone": phone,
        "city": city,
        "companies": company,
        "latest_company": company,
        "job_titles": job,
        "latest_role": job,
        "degree": education,
        "experience_duration": exp,
        "has_work_experience": _has_exp(exp),
        "applied_at": applied,
        "status": _status(status_raw, feedback),
        "notes": "\n".join(notes),
        "tags": ["bde", "business_development", "excel_import", "apna"],
    }


async def ensure_role() -> None:
    async with async_session() as db:
        r = await db.get(HiringRole, ROLE_ID)
        if not r:
            db.add(
                HiringRole(
                    id=ROLE_ID,
                    name=ROLE_NAME,
                    description="Business Development Executive applicants (Naukri/Apna/BDE exports)",
                    is_active=True,
                    sort_order=20,
                )
            )
            await db.commit()
            print(f"Created role {ROLE_ID}")
        else:
            print(f"Role exists: {ROLE_ID}")


async def run(files: list[Path], reset: bool) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    await ensure_role()

    if reset:
        async with async_session() as db:
            old = (
                await db.execute(select(Candidate).where(Candidate.role_id == ROLE_ID))
            ).scalars().all()
            for c in old:
                await db.delete(c)
            await db.commit()
            print(f"Cleared {len(old)} previous BDE candidates")

    by_id: dict[str, dict] = {}
    for path in files:
        if not path.exists():
            print(f"MISSING {path}")
            continue
        try:
            sheets = load_xlsx_all_sheets(path)
        except Exception as e:
            print(f"ERROR {path.name}: {e}")
            continue
        file_rows = 0
        for sn, rows in sheets:
            for row in rows:
                payload = row_to_payload(row, f"{path.name}#{sn}")
                if not payload:
                    continue
                cid = payload["id"]
                if cid in by_id:
                    prev = by_id[cid]
                    for k, v in payload.items():
                        if k == "id":
                            continue
                        if k == "notes" and v and v not in (prev.get("notes") or ""):
                            prev["notes"] = ((prev.get("notes") or "") + "\n" + v).strip()
                        elif v and not prev.get(k):
                            prev[k] = v
                else:
                    by_id[cid] = payload
                file_rows += 1
        print(f"{path.name}: rows={file_rows} unique_so_far={len(by_id)}")

    created = 0
    updated = 0
    now = datetime.now(timezone.utc).replace(tzinfo=None)

    async with async_session() as db:
        for payload in by_id.values():
            cid = payload["id"]
            existing = await db.get(Candidate, cid)
            fields = {
                "role_id": ROLE_ID,
                "role_name": ROLE_NAME,
                "status": payload.get("status") or "new",
                "tags": payload.get("tags") or ["bde"],
                "notes": payload.get("notes") or "",
                "name": payload.get("name"),
                "email": payload.get("email"),
                "phone": payload.get("phone"),
                "city": payload.get("city"),
                "companies": payload.get("companies"),
                "latest_company": payload.get("latest_company"),
                "job_titles": payload.get("job_titles"),
                "latest_role": payload.get("latest_role"),
                "other_skills": payload.get("other_skills"),
                "career_objective": payload.get("career_objective"),
                "work_experience_detail": payload.get("work_experience_detail"),
                "degree": payload.get("degree"),
                "institute": payload.get("institute"),
                "graduation_year": payload.get("graduation_year"),
                "experience_duration": payload.get("experience_duration"),
                "has_work_experience": payload.get("has_work_experience"),
                "availability": payload.get("availability"),
                "applied_at": payload.get("applied_at"),
                "updated_at": now,
            }
            if existing:
                for k, v in fields.items():
                    if v is not None and v != "":
                        setattr(existing, k, v)
                updated += 1
            else:
                db.add(Candidate(id=cid, created_at=now, starred=False, **fields))
                created += 1
        await db.commit()

    async with async_session() as db:
        total = (
            await db.execute(
                select(func.count()).select_from(Candidate).where(Candidate.role_id == ROLE_ID)
            )
        ).scalar()
    print("\n=== BDE import done ===")
    print(f"unique: {len(by_id)}  created: {created}  updated: {updated}")
    print(f"role {ROLE_NAME}: {total}")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--reset", action="store_true")
    args = p.parse_args()
    files = [f for f in DEFAULT_FILES if f.exists()]
    if not files:
        raise SystemExit("No BDE files found")
    asyncio.run(run(files, args.reset))


if __name__ == "__main__":
    main()
