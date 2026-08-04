"""
Import:
  1) Internshala scraped internship listings → jobs table
  2) Interviews.xlsx → hiring candidates

Usage (from BE/):
  python -m app.import_internshala_listings
  python -m app.import_internshala_listings --reset-jobs
"""

from __future__ import annotations

import argparse
import asyncio
import csv
import hashlib
import re
import sys
import uuid
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

import openpyxl
from sqlalchemy import select, func, delete

from app.database import async_session, engine, Base
from app.models.job import Job
from app.models.hiring import Candidate, HiringRole

DOWNLOADS = Path.home() / "Downloads"

# Prefer CSV (fast) over sibling .xlsx; keep unique workbook formats.
LISTING_FILES = [
    DOWNLOADS / "internshala-internships.csv",
    DOWNLOADS / "internshala-internships2.csv",
    DOWNLOADS / "internshala-internships2 (1).xlsx",  # alternate column layout
    DOWNLOADS / "internshala-internships3.csv",
    DOWNLOADS / "internshala-internships4.csv",
]
# Full set if CSVs missing (legacy):
LISTING_FILES_FALLBACK = [
    DOWNLOADS / "internshala-internships.xlsx",
    DOWNLOADS / "internshala-internships2.xlsx",
    DOWNLOADS / "internshala-internships3.xlsx",
    DOWNLOADS / "internshala-internships4.xlsx",
]

INTERVIEWS_FILE = DOWNLOADS / "Interviews.xlsx"

TAG_PREFIX = "is_listing_"  # posted_by tag for reset


def _norm_headers(row: dict) -> dict:
    return {str(k).lstrip("\ufeff").strip(): v for k, v in row.items() if k is not None}


def _get(row: dict, *keys: str):
    lower = {str(k).lower(): v for k, v in row.items()}
    for key in keys:
        if key.lower() in lower and lower[key.lower()] is not None:
            s = str(lower[key.lower()]).strip()
            if s and s.lower() not in {"none", "nan", "null"}:
                return s
    return None


def _phone(raw: str | None) -> str | None:
    if not raw:
        return None
    d = re.sub(r"\D", "", str(raw).split(".")[0] if "." in str(raw) else str(raw))
    return d[-10:] if len(d) >= 10 else (d or None)


def _email(raw: str | None) -> str | None:
    if not raw:
        return None
    s = str(raw).strip()
    return s.lower() if "@" in s else None


def parse_stipend(raw: str | None) -> dict | None:
    if not raw:
        return None
    # "10000 /month", "4500-9000 /month", "Performance Based"
    m = re.search(r"(\d+)\s*[-–]\s*(\d+)", raw.replace(",", ""))
    if m:
        return {"amount": int(m.group(1)), "currency": "INR", "period": "monthly", "max": int(m.group(2)), "raw": raw}
    m = re.search(r"(\d+)", raw.replace(",", ""))
    if m:
        return {"amount": int(m.group(1)), "currency": "INR", "period": "monthly", "raw": raw}
    return {"amount": 0, "currency": "INR", "period": "monthly", "raw": raw}


def category_from_title(title: str) -> str:
    t = title.lower()
    if any(x in t for x in ("web", "react", "front end", "frontend", "full stack", "android", "ios", "app development")):
        return "Software Development"
    if any(x in t for x in ("market", "sales", "business development", "seo", "social media", "content")):
        return "Marketing"
    if any(x in t for x in ("design", "graphic", "ui", "ux")):
        return "Design"
    if any(x in t for x in ("data", "ml", "ai", "analyst")):
        return "Data Science"
    if any(x in t for x in ("hr", "human resource", "recruit")):
        return "Human Resources"
    return "Other"


def listing_from_row(row: dict, source: str) -> dict | None:
    row = _norm_headers(row)
    # Format A: scraper columns
    title = _get(row, "view_detail_button", "Role", "Web Development")  # broken xlsx may put title in first col
    link = _get(row, "view_detail_button href", "Job Link")
    company = _get(row, "link_display_like_text", "Company")
    company_link = _get(row, "link_display_like_text href", "Company Link")
    logo = _get(row, "internship_logo src", "Logo")
    location = _get(row, "location_link", "Job Type")
    duration = _get(row, "item_body", "Duration")
    stipend_raw = _get(row, "stipend", "Stipend")
    posted = _get(row, "item_body 2", "Posted Date")
    # Format B: clean Role/Job Link headers (internships2 (1).xlsx)
    if not title:
        title = _get(row, "Role")
    if not link:
        link = _get(row, "Job Link")
    if not company:
        company = _get(row, "Company")
    if not location:
        location = _get(row, "Job Type", "location_link")
    if not duration:
        # sometimes duration is item_body
        duration = _get(row, "Duration", "item_body")

    # Broken xlsx where first row became headers (title is a random internship name as header)
    # If keys look like URLs/titles, reconstruct from first values... skip broken sheets without proper headers
    if not title and not company:
        # try positional: if keys aren't standard, row values might still work if we used wrong header
        vals = list(row.values())
        keys = list(row.keys())
        # if first key looks like a role title (not a header name)
        if keys and keys[0] and "http" not in keys[0].lower() and keys[0].lower() not in {
            "view_detail_button", "role", "name"
        }:
            title = keys[0]
            # values: link, company, company_link, logo, location, ...
            if len(vals) >= 1 and vals[0] and "http" in str(vals[0]):
                link = str(vals[0])
            if len(vals) >= 2:
                company = str(vals[1]) if vals[1] else company
            if len(vals) >= 5 and vals[4]:
                location = str(vals[4])
            if len(vals) >= 9 and vals[8]:
                duration = str(vals[8])
            if len(vals) >= 10 and vals[9]:
                stipend_raw = str(vals[9])

    if not title:
        return None
    if title.lower() in {"view_detail_button", "role", "name"}:
        return None

    company = company or "Unknown"
    location = location or "Not specified"
    is_remote = "work from home" in location.lower() or "remote" in location.lower() or "wfh" in location.lower()

    key = link or f"{title}|{company}|{location}|{duration}"
    jid = "isjob_" + hashlib.sha1(key.encode()).hexdigest()[:16]

    stipend = parse_stipend(stipend_raw)
    desc_parts = [
        f"Internship listing imported from Internshala scrape ({source}).",
        f"Title: {title}",
        f"Company: {company}",
        f"Location: {location}",
    ]
    if duration:
        desc_parts.append(f"Duration: {duration}")
    if stipend_raw:
        desc_parts.append(f"Stipend: {stipend_raw}")
    if posted:
        desc_parts.append(f"Posted: {posted}")
    if link:
        desc_parts.append(f"Apply: {link}")
    if company_link:
        desc_parts.append(f"Company page: {company_link}")

    deadline = posted or "Open"
    # try parse apply by date from item_body 2 like "31 Jul' 22"
    if posted and re.search(r"\d{1,2}\s+\w+", posted):
        deadline = posted

    return {
        "id": jid,
        "title": title[:200],
        "company": company[:200],
        "company_logo": logo,
        "location": location[:200],
        "is_remote": is_remote,
        "type": "internship",
        "category": category_from_title(title),
        "description": "\n".join(desc_parts),
        "requirements": [],
        "responsibilities": [],
        "salary": None,
        "stipend": stipend,
        "duration": duration,
        "application_deadline": str(deadline)[:80],
        "posted_by": TAG_PREFIX + "import",
        "status": "approved",
        "applicants": [],
        "source_link": link,
    }


def load_csv(path: Path) -> list[dict]:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        return list(csv.DictReader(fh))


def load_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows_out: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        try:
            headers = next(it)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(headers)]
        for raw in it:
            if not raw or all(c is None or str(c).strip() == "" for c in raw):
                continue
            rows_out.append({headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))})
    wb.close()
    return rows_out


async def import_listings(reset: bool) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    if reset:
        async with async_session() as db:
            res = await db.execute(delete(Job).where(Job.posted_by.like(f"{TAG_PREFIX}%")))
            await db.commit()
            print(f"Cleared previous Internshala listing jobs (delete ok)")

    paths = [p for p in LISTING_FILES if p.exists()]
    if not paths:
        paths = [p for p in LISTING_FILES_FALLBACK if p.exists()]
    if not paths:
        print("No listing files found in Downloads")
        return

    by_id: dict[str, dict] = {}
    for path in paths:
        print(f"Loading {path.name}…", flush=True)
        try:
            rows = load_xlsx(path) if path.suffix.lower() == ".xlsx" else load_csv(path)
        except Exception as e:
            print(f"ERROR {path.name}: {e}", flush=True)
            continue
        n = 0
        for row in rows:
            payload = listing_from_row(row, path.name)
            if not payload:
                continue
            jid = payload["id"]
            if jid not in by_id:
                by_id[jid] = payload
                n += 1
            else:
                prev = by_id[jid]
                for k, v in payload.items():
                    if v and not prev.get(k):
                        prev[k] = v
        print(f"{path.name}: rows={len(rows)} new_unique+={n} total_unique={len(by_id)}", flush=True)

    print(f"Upserting {len(by_id)} jobs…", flush=True)
    created = updated = 0
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    ids = list(by_id.keys())
    existing_ids: set[str] = set()
    async with async_session() as db:
        # Load existing ids in chunks
        for i in range(0, len(ids), 500):
            chunk = ids[i : i + 500]
            res = await db.execute(select(Job.id).where(Job.id.in_(chunk)))
            existing_ids.update(r[0] for r in res.all())

        batch: list = []
        BATCH = 200
        for jid, payload in by_id.items():
            if jid in existing_ids:
                existing = await db.get(Job, jid)
                if existing:
                    existing.title = payload["title"]
                    existing.company = payload["company"]
                    existing.company_logo = payload.get("company_logo")
                    existing.location = payload["location"]
                    existing.is_remote = payload["is_remote"]
                    existing.duration = payload.get("duration")
                    existing.stipend = payload.get("stipend")
                    existing.description = payload["description"]
                    existing.status = "approved"
                    updated += 1
            else:
                batch.append(
                    Job(
                        id=jid,
                        title=payload["title"],
                        company=payload["company"],
                        company_logo=payload.get("company_logo"),
                        location=payload["location"],
                        is_remote=payload["is_remote"],
                        type="internship",
                        category=payload["category"],
                        description=payload["description"],
                        requirements=[],
                        responsibilities=[],
                        salary=None,
                        stipend=payload.get("stipend"),
                        duration=payload.get("duration"),
                        application_deadline=payload.get("application_deadline") or "Open",
                        posted_by=payload["posted_by"],
                        posted_at=now,
                        status="approved",
                        applicants=[],
                    )
                )
                created += 1
                if len(batch) >= BATCH:
                    db.add_all(batch)
                    await db.commit()
                    print(f"  committed batch created={created} updated={updated}", flush=True)
                    batch = []
        if batch:
            db.add_all(batch)
        await db.commit()

    async with async_session() as db:
        total = (
            await db.execute(select(func.count()).select_from(Job).where(Job.type == "internship"))
        ).scalar()
        listed = (
            await db.execute(
                select(func.count()).select_from(Job).where(Job.posted_by.like(f"{TAG_PREFIX}%"))
            )
        ).scalar()
    print(f"\nListings: unique={len(by_id)} created={created} updated={updated}", flush=True)
    print(f"Jobs internships total={total}  is_listing import={listed}", flush=True)


async def import_interviews() -> None:
    path = INTERVIEWS_FILE
    if not path.exists():
        print("MISSING Interviews.xlsx")
        return

    async with async_session() as db:
        if not await db.get(HiringRole, "interview_pipeline"):
            db.add(
                HiringRole(
                    id="interview_pipeline",
                    name="Interview Pipeline",
                    description="Candidates from Interviews.xlsx tracker",
                    is_active=True,
                    sort_order=15,
                )
            )
            await db.commit()

    rows = load_xlsx(path)
    created = updated = 0
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    async with async_session() as db:
        for row in rows:
            row = _norm_headers(row)
            name = _get(row, "Name")
            phone = _phone(_get(row, "Phone No", "Phone", "Phone Number"))
            email = _email(_get(row, "Email", "Email ID"))
            if not name and not phone:
                continue
            role_applied = _get(row, "Role") or "Intern"
            education = _get(row, "Education")
            details = _get(row, "Details")
            status_raw = _get(row, "Status")
            comments = _get(row, "Comments")
            joining = _get(row, "Joining Date")
            contacted = _get(row, "Contacted")

            key = email or f"{phone or ''}|{(name or '').lower()}"
            cid = "iv_" + hashlib.sha1(key.encode()).hexdigest()[:12]
            notes = ["Source: Interviews.xlsx"]
            if role_applied:
                notes.append(f"Applied role: {role_applied}")
            if details:
                notes.append(f"Details: {details}")
            if comments:
                notes.append(f"Comments: {comments}")
            if contacted:
                notes.append(f"Contacted: {contacted}")
            if joining:
                notes.append(f"Joining: {joining}")

            st = "interview"
            blob = f"{status_raw or ''} {comments or ''}".lower()
            if "reject" in blob or "not interested" in blob:
                st = "rejected"
            elif "select" in blob or "join" in blob or "offer" in blob:
                st = "offer"
            elif "hold" in blob or "looking" in blob:
                st = "on_hold"

            fields = dict(
                role_id="interview_pipeline",
                role_name="Interview Pipeline",
                status=st,
                tags=["interviews_xlsx", "import"],
                notes="\n".join(notes),
                name=name.strip() if name else None,
                phone=phone,
                email=email,
                degree=education,
                experience_duration=details,
                has_work_experience="Yes" if details and re.search(r"(?i)exp|year|month", details) else None,
                latest_role=role_applied,
                updated_at=now,
            )
            existing = await db.get(Candidate, cid)
            if existing:
                for k, v in fields.items():
                    if v is not None and v != "":
                        setattr(existing, k, v)
                updated += 1
            else:
                db.add(Candidate(id=cid, created_at=now, starred=False, **fields))
                created += 1
        await db.commit()
    print(f"Interviews.xlsx: created={created} updated={updated}")


async def main_async(reset_jobs: bool) -> None:
    await import_listings(reset_jobs)
    await import_interviews()


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--reset-jobs", action="store_true")
    args = p.parse_args()
    asyncio.run(main_async(args.reset_jobs))


if __name__ == "__main__":
    main()
