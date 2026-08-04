"""
Import Digital Marketing + Executive/Sr Executive Excel exports.

Usage (from BE/):
  python -m app.import_marketing_executive
  python -m app.import_marketing_executive --reset
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

import openpyxl
from sqlalchemy import select, func

from app.database import async_session, engine, Base
from app.models.hiring import Candidate, HiringRole

DOWNLOADS = Path.home() / "Downloads"

# (path, role_id, role_name)
FILE_SPECS: list[tuple[Path, str, str]] = [
    (
        DOWNLOADS / "digital_marketing_manager_20_02_2023_at_12_32_38_pm.xlsx",
        "content_social_media",
        "Content & Social Media Marketing",
    ),
    (
        DOWNLOADS / "DigitalMarketing_645464965_12-16-2021.xlsx",
        "content_social_media",
        "Content & Social Media Marketing",
    ),
    (
        DOWNLOADS / "Executive-Sr.-Execut_20201207194557_160.xlsx",
        "management_executive",
        "Executive / Sr. Executive / Management",
    ),
    (
        DOWNLOADS / "Executive-Sr.-Execut_20201207194641_160.xlsx",
        "management_executive",
        "Executive / Sr. Executive / Management",
    ),
    (
        DOWNLOADS / "Executive-Sr.-Execut_20201207194658_160.xlsx",
        "management_executive",
        "Executive / Sr. Executive / Management",
    ),
    (
        DOWNLOADS / "Executive-Sr.-Execut_20201207194720_24.xlsx",
        "management_executive",
        "Executive / Sr. Executive / Management",
    ),
]

ID_PREFIX = {
    "content_social_media": "dmkt",
    "management_executive": "exec",
}


def _get(row: dict, *keys: str):
    lower = {str(k).lstrip("\ufeff").strip().lower(): v for k, v in row.items() if k is not None}
    for key in keys:
        if key.lower() in lower:
            v = lower[key.lower()]
            if v is None:
                continue
            s = str(v).strip()
            if re.fullmatch(r"\d+\.0", s):
                s = s[:-2]
            if s and s.lower() not in {"none", "nan", "null", "na", "n/a"}:
                return s
    return None


def _phone(raw: str | None) -> str | None:
    if not raw:
        return None
    for p in re.split(r"[,;/]", raw):
        d = re.sub(r"\D", "", p)
        if len(d) >= 10:
            return d[-10:]
    d = re.sub(r"\D", "", raw)
    return d[-10:] if len(d) >= 10 else (d or None)


def _email(raw: str | None) -> str | None:
    if not raw:
        return None
    for p in re.split(r"[,;\s]+", raw):
        if "@" in p:
            return p.strip().lower()
    return raw.strip().lower() or None


def _has_exp(exp: str | None) -> str | None:
    if not exp:
        return None
    if re.search(r"(?i)fresher|0\s*year", exp) and not re.search(r"(?i)[1-9]", exp):
        return "No"
    if re.search(r"(?i)\d|<1", exp):
        return "Yes"
    return None


def _status(*parts: str | None) -> str:
    blob = " ".join(p for p in parts if p).lower()
    if "reject" in blob or "not interested" in blob:
        return "rejected"
    if "shortlist" in blob or "selected" in blob:
        return "shortlisted"
    if "interview" in blob:
        return "interview"
    if any(x in blob for x in ("hold", "pending", "callback", "call back", "not answering", "not connected")):
        return "on_hold"
    return "new"


def load_sheets(path: Path) -> list[tuple[str, list[dict]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple[str, list[dict]]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
        rows = []
        for raw in it:
            if not raw or all(c is None or str(c).strip() == "" for c in raw):
                continue
            rows.append({headers[i]: (raw[i] if i < len(raw) else None) for i in range(len(headers))})
        if rows:
            out.append((sn, rows))
    wb.close()
    return out


def row_to_payload(row: dict, source: str, role_id: str, role_name: str) -> dict | None:
    name = _get(row, "Name")
    email = _email(_get(row, "Email ID", "Email"))
    phone = _phone(_get(row, "Phone Number", "Phone"))
    if not name and not email and not phone:
        return None

    # Naukri-ish
    if _get(row, "Email ID") or _get(row, "Job Title") or _get(row, "Current Location") or _get(row, "Requirement Name"):
        city = _get(row, "Current Location", "City")
        exp = _get(row, "Total Experience", "Experience")
        company = _get(row, "Curr. Company name", "Company")
        designation = _get(row, "Curr. Company Designation", "Current Job", "Role")
        skills = _get(row, "Key Skills", "Skills")
        summary = _get(row, "Summary")
        headline = _get(row, "Resume Headline")
        ug = _get(row, "Under Graduation degree")
        ug_inst = _get(row, "UG University/institute Name")
        ug_year = _get(row, "UG Graduation year")
        availability = _get(row, "Notice period/ Availability to join", "Notice Period")
        applied = _get(row, "Date of application", "Applied On")
        notes_extra = _get(row, "Notes")
        feedback = _get(row, "Feedback")
        status_raw = _get(row, "Status")
        job_title = _get(row, "Job Title", "Requirement Name")
        req_id = _get(row, "Requirement Id")
        source_app = _get(row, "Source Of Application")
        functional = _get(row, "Functional Area")
        industry = _get(row, "Industry")

        notes = [f"Source: {source}"]
        if job_title:
            notes.append(f"Job: {job_title}")
        if req_id:
            notes.append(f"Requirement: {req_id}")
        if source_app:
            notes.append(f"App source: {source_app}")
        if notes_extra:
            notes.append(f"Notes: {notes_extra}")
        if feedback:
            notes.append(f"Feedback: {feedback}")
        if status_raw:
            notes.append(f"Export status: {status_raw}")
        if functional:
            notes.append(f"Functional area: {functional}")
        if industry:
            notes.append(f"Industry: {industry}")

        prefix = ID_PREFIX.get(role_id, "imp")
        key = email or f"{phone or ''}|{(name or '').lower()}"
        cid = f"{prefix}_{hashlib.sha1(key.encode()).hexdigest()[:12]}"

        return dict(
            id=cid,
            role_id=role_id,
            role_name=role_name,
            status=_status(status_raw, notes_extra, feedback),
            tags=["excel_import", role_id],
            notes="\n".join(notes),
            name=name,
            email=email,
            phone=phone,
            city=city.split(",")[0].strip() if city else None,
            experience_duration=exp,
            has_work_experience=_has_exp(exp),
            companies=company,
            latest_company=company,
            job_titles=designation,
            latest_role=designation,
            other_skills=skills,
            career_objective=headline or summary,
            work_experience_detail=summary,
            degree=ug,
            institute=ug_inst,
            graduation_year=ug_year,
            availability=availability,
            applied_at=applied,
        )

    # Apna-style
    city = _get(row, "Area", "Current Location")
    company = _get(row, "Company")
    job = _get(row, "Current Job")
    education = _get(row, "Education")
    exp = _get(row, "Experience", "Total Experience")
    status_raw = _get(row, "Status")
    feedback = _get(row, "Feedback")
    applied = _get(row, "Applied On", "Date of application")

    prefix = ID_PREFIX.get(role_id, "imp")
    key = f"{phone or ''}|{(name or '').lower()}"
    cid = f"{prefix}_{hashlib.sha1(key.encode()).hexdigest()[:12]}"
    notes = [f"Source: {source}"]
    if feedback:
        notes.append(f"Feedback: {feedback}")
    if status_raw:
        notes.append(f"Status: {status_raw}")

    return dict(
        id=cid,
        role_id=role_id,
        role_name=role_name,
        status=_status(status_raw, feedback),
        tags=["excel_import", "apna", role_id],
        notes="\n".join(notes),
        name=name,
        email=email,
        phone=phone,
        city=city,
        companies=company,
        latest_company=company,
        job_titles=job,
        latest_role=job,
        degree=education,
        experience_duration=exp,
        has_work_experience=_has_exp(exp),
        applied_at=applied,
    )


async def ensure_role(role_id: str, role_name: str, sort_order: int) -> None:
    async with async_session() as db:
        r = await db.get(HiringRole, role_id)
        if not r:
            db.add(
                HiringRole(
                    id=role_id,
                    name=role_name,
                    description="Imported from Excel applications",
                    is_active=True,
                    sort_order=sort_order,
                )
            )
            await db.commit()
            print(f"Created role {role_id}")
        else:
            # keep existing name for content_social_media
            if role_id == "management_executive" and r.name != role_name:
                r.name = role_name
                await db.commit()
            print(f"Role exists: {role_id}")


async def run(reset: bool) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    roles_needed = {(rid, rname) for _, rid, rname in FILE_SPECS}
    for i, (rid, rname) in enumerate(sorted(roles_needed)):
        await ensure_role(rid, rname, 30 + i)

    role_ids = {rid for _, rid, _ in FILE_SPECS}
    if reset:
        async with async_session() as db:
            # only delete candidates from these imports (id prefixes)
            old = (await db.execute(select(Candidate).where(Candidate.role_id.in_(role_ids)))).scalars().all()
            n = 0
            for c in old:
                if (c.id or "").startswith(("dmkt_", "exec_")) or "excel_import" in (c.tags or []):
                    # only remove exec_ / dmkt_ to avoid wiping prior content_social_media
                    if (c.id or "").startswith(("dmkt_", "exec_")):
                        await db.delete(c)
                        n += 1
            await db.commit()
            print(f"Cleared {n} previous dmkt_/exec_ candidates")

    by_id: dict[str, dict] = {}
    for path, role_id, role_name in FILE_SPECS:
        if not path.exists():
            print(f"MISSING {path}")
            continue
        try:
            sheets = load_sheets(path)
        except Exception as e:
            print(f"ERROR {path.name}: {e}")
            continue
        n = 0
        for sn, rows in sheets:
            for row in rows:
                payload = row_to_payload(row, f"{path.name}#{sn}", role_id, role_name)
                if not payload:
                    continue
                cid = payload["id"]
                if cid in by_id:
                    prev = by_id[cid]
                    # prefer more specific later file for missing fields
                    for k, v in payload.items():
                        if k == "id":
                            continue
                        if k == "notes" and v and v not in (prev.get("notes") or ""):
                            prev["notes"] = ((prev.get("notes") or "") + "\n" + v).strip()
                        elif v and not prev.get(k):
                            prev[k] = v
                    # if same person appears in marketing vs executive, keep first role unless exec is clearer
                else:
                    by_id[cid] = payload
                n += 1
        print(f"{path.name}: rows={n} unique_so_far={len(by_id)} → {role_id}")

    created = updated = 0
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    async with async_session() as db:
        for payload in by_id.values():
            cid = payload["id"]
            existing = await db.get(Candidate, cid)
            fields = {
                "role_id": payload["role_id"],
                "role_name": payload["role_name"],
                "status": payload.get("status") or "new",
                "tags": payload.get("tags") or [],
                "notes": payload.get("notes") or "",
                "name": payload.get("name"),
                "email": payload.get("email"),
                "phone": payload.get("phone"),
                "city": payload.get("city"),
                "companies": payload.get("companies"),
                "latest_company": payload.get("latest_company"),
                "job_titles": payload.get("job_titles"),
                "latest_role": payload.get("latest_role"),
                "other_skills": payload.get("other_skills"),
                "career_objective": payload.get("career_objective"),
                "work_experience_detail": payload.get("work_experience_detail"),
                "degree": payload.get("degree"),
                "institute": payload.get("institute"),
                "graduation_year": payload.get("graduation_year"),
                "experience_duration": payload.get("experience_duration"),
                "has_work_experience": payload.get("has_work_experience"),
                "availability": payload.get("availability"),
                "applied_at": payload.get("applied_at"),
                "updated_at": now,
            }
            if existing:
                for k, v in fields.items():
                    if v is not None and v != "":
                        setattr(existing, k, v)
                updated += 1
            else:
                db.add(Candidate(id=cid, created_at=now, starred=False, **fields))
                created += 1
        await db.commit()

    async with async_session() as db:
        print("\n=== Marketing / Executive import done ===")
        print(f"unique: {len(by_id)}  created: {created}  updated: {updated}")
        for rid in sorted(role_ids):
            # count only our prefixes for content_social_media
            if rid == "content_social_media":
                cnt = (
                    await db.execute(
                        select(func.count())
                        .select_from(Candidate)
                        .where(Candidate.id.like("dmkt_%"))
                    )
                ).scalar()
                total_role = (
                    await db.execute(
                        select(func.count())
                        .select_from(Candidate)
                        .where(Candidate.role_id == rid)
                    )
                ).scalar()
                print(f"  {rid}: +dmkt_={cnt}  total_in_role={total_role}")
            else:
                cnt = (
                    await db.execute(
                        select(func.count()).select_from(Candidate).where(Candidate.role_id == rid)
                    )
                ).scalar()
                print(f"  {rid}: {cnt}")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--reset", action="store_true", help="Delete previous dmkt_/exec_ imports first")
    args = p.parse_args()
    asyncio.run(run(args.reset))


if __name__ == "__main__":
    main()
