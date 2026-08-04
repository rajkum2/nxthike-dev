"""
Import Internshala application Excel exports + intern lists.

Usage (from BE/):
  python -m app.import_internshala_excel
  python -m app.import_internshala_excel --reset
"""

from __future__ import annotations

import argparse
import asyncio
import hashlib
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from dotenv import load_dotenv

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

import openpyxl
from sqlalchemy import select, func

from app.database import async_session, engine, Base
from app.models.hiring import Candidate, HiringRole

DOWNLOADS = Path.home() / "Downloads"

# filename → role
ROLE_PATTERNS: list[tuple[re.Pattern[str], str, str]] = [
    (re.compile(r"AI_Agent_Development|AI Agent", re.I), "ai_agent_development", "AI Agent Development"),
    (re.compile(r"Web_Development|Web Development", re.I), "web_development_internship", "Web Development Internship"),
    (re.compile(r"Data_Entry|Data Entry", re.I), "data_entry", "Data Entry Internship"),
    (re.compile(r"Telecalling", re.I), "telecaller", "Telecaller / Outbound"),
    (re.compile(r"Int-Student|Interns List", re.I), "internship_general", "Internship (General)"),
]

DEFAULT_ROLE = ("internship_general", "Internship (General)")

FILES = [
    DOWNLOADS / "Int-Student-1.xlsx",
    DOWNLOADS / "Interns List.xlsx",
    DOWNLOADS / "Internshala_open_Applications_for_Data_Entry_internship_2022_05_21_10_10_06 (1).xlsx",
    DOWNLOADS / "Internshala_open_Applications_for_Data_Entry_internship_2022_05_21_10_10_06 (1).xlsx - Sheet1 (1).xlsx",
    DOWNLOADS / "Internshala_open_Applications_for_Data_Entry_internship_2022_05_21_10_10_06 (1).xlsx - Sheet1.xlsx",
    DOWNLOADS / "Internshala_open_Applications_for_Data_Entry_internship_2022_11_03_10_47_06.xlsx",
    DOWNLOADS / "Internshala_open_Applications_for_Data_Entry_internship_2023_06_10_07_06_03.xlsx",
    DOWNLOADS / "Internshala_open_Applications_for_Web_Development_internship_2021_10_28_12_30_09.xlsx",
    DOWNLOADS / "Internshala_shortlisted_Applications_for_AI_Agent_Development_internship_2025_07_17_23_13_05.xlsx",
    DOWNLOADS / "Internshala_shortlisted_Applications_for_Data_Entry_internship_2022_05_22_21_31_05.xlsx",
    DOWNLOADS / "Internshala_shortlisted_Applications_for_Telecalling_internship_2022_06_15_12_22_06.xlsx",
    DOWNLOADS / "Internshala_shortlisted_Applications_for_Telecalling_internship_2022_06_15_12_28_05.xlsx",
    DOWNLOADS / "Internshala_shortlisted_Applications_for_Telecalling_internship_2022_10_25_11_15_07.xlsx",
    DOWNLOADS / "Internshala_shortlisted_Applications_for_Telecalling_internship_2022_10_25_12_01_07.xlsx",
]


def infer_role(filename: str, applied_for: str | None = None) -> tuple[str, str]:
    blob = f"{filename} {applied_for or ''}"
    for pat, rid, rname in ROLE_PATTERNS:
        if pat.search(blob):
            return rid, rname
    # applied for column
    if applied_for:
        af = applied_for.lower()
        if "telecall" in af:
            return "telecaller", "Telecaller / Outbound"
        if "data entry" in af:
            return "data_entry", "Data Entry Internship"
        if "web" in af:
            return "web_development_internship", "Web Development Internship"
        if "ai agent" in af or "ai/ml" in af:
            return "ai_agent_development", "AI Agent Development"
    return DEFAULT_ROLE


def _get(row: dict, *keys: str):
    lower = {str(k).strip().lower(): v for k, v in row.items() if k is not None and str(k).strip()}
    for key in keys:
        if key.lower() in lower:
            v = lower[key.lower()]
            if v is None:
                continue
            s = str(v).strip()
            if re.fullmatch(r"\d+\.0", s):
                s = s[:-2]
            if s and s.lower() not in {"none", "nan", "null", "na"}:
                return s
    return None


def _phone(raw: str | None) -> str | None:
    if not raw:
        return None
    d = re.sub(r"\D", "", raw)
    if len(d) >= 10:
        return d[-10:]
    return d or None


def _skill_scores(row: dict) -> list[str]:
    """Collect 'Skill (out of 3)' columns where score > 0."""
    skills = []
    for k, v in row.items():
        if k is None:
            continue
        ks = str(k)
        m = re.search(r"^(.+?)\s*\(out of 3\)\s*$", ks, re.I)
        if not m:
            continue
        skill = m.group(1).strip()
        try:
            score = float(v) if v is not None and str(v).strip() != "" else 0
        except (TypeError, ValueError):
            score = 0
        if score > 0:
            skills.append(f"{skill}:{int(score) if score == int(score) else score}/3")
    return skills


def _status(*parts: str | None, shortlisted_file: bool = False) -> str:
    blob = " ".join(p for p in parts if p).lower()
    if shortlisted_file or "shortlist" in blob:
        return "shortlisted"
    if "reject" in blob:
        return "rejected"
    if "interview" in blob:
        return "interview"
    if "hold" in blob or "pending" in blob or "contacted" in blob:
        return "on_hold"
    return "new"


def load_sheets(path: Path) -> list[tuple[str, list[dict]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: list[tuple[str, list[dict]]] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        it = ws.iter_rows(values_only=True)
        try:
            header_row = next(it)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(header_row)]
        # drop completely empty header names trailing
        rows = []
        for raw in it:
            if not raw or all(c is None or str(c).strip() == "" for c in raw):
                continue
            d = {}
            for i, h in enumerate(headers):
                if not h or h.startswith("col"):
                    continue
                d[h] = raw[i] if i < len(raw) else None
            if any(v is not None and str(v).strip() for v in d.values()):
                rows.append(d)
        if rows:
            out.append((sn, rows))
    wb.close()
    return out


def row_to_payload(row: dict, source: str, shortlisted: bool) -> dict | None:
    name = _get(row, "Name")
    phone = _phone(_get(row, "Phone", "Phone Number"))
    email = _get(row, "Email Address", "Email", "Email ID")
    if email:
        email = email.split(",")[0].strip().lower()
    if not name and not phone and not email:
        return None

    applied_for = _get(row, "Applied for", "Applying for")
    role_id, role_name = infer_role(source, applied_for)

    city = _get(row, "Current City", "City", "Location")
    institute = _get(row, "Institute", "College")
    degree = _get(row, "Degree", "Education")
    stream = _get(row, "Stream")
    year = _get(row, "Current Year Of Graduation", "Current Year")
    other_skills = _get(row, "Other skills", "Other Skills", "Skills")
    skill_scores = _skill_scores(row)
    if skill_scores:
        scored = ", ".join(skill_scores)
        other_skills = f"{other_skills}; {scored}" if other_skills else scored

    why = _get(
        row,
        "Why should you be hired for this role?",
        "Q1. Why should you be hired for this rol",
        "Q1. Why should you be hired for this role?",
    )
    availability = _get(
        row,
        "Are you available for 3 months, starting",
        "Are you available for 2 months, starting",
        "Q2. Are you available for 3 months, star",
        "Q2. Are you available for 3 months, starting",
        "Availability",
    )
    app_link = _get(row, "Link to application", "Application Link")
    notes_col = _get(row, "Notes", "Interview Notes")
    stage = _get(row, "Stage", "Status")
    perf_ug = _get(row, "Performance_UG")
    perf_pg = _get(row, "Performance_PG")
    perf_12 = _get(row, "Performance_12")

    notes = [f"Source: {source}"]
    if applied_for:
        notes.append(f"Applied for: {applied_for}")
    if stage:
        notes.append(f"Stage/Status: {stage}")
    if notes_col:
        notes.append(f"Notes: {notes_col}")
    if why:
        notes.append(f"Why hire: {why[:500]}")

    key = (email or f"{phone or ''}|{(name or '').lower()}|{role_id}")
    # allow same person in multiple roles
    cid = f"is_{hashlib.sha1(key.encode()).hexdigest()[:12]}"

    return {
        "id": cid,
        "role_id": role_id,
        "role_name": role_name,
        "status": _status(stage, notes_col, shortlisted_file=shortlisted),
        "tags": ["internshala", "excel_import", role_id] + (["shortlisted_export"] if shortlisted else ["open_export"]),
        "notes": "\n".join(notes),
        "name": name.strip() if name else None,
        "phone": phone,
        "email": email,
        "city": city,
        "institute": institute,
        "degree": degree,
        "stream": stream,
        "graduation_year": year,
        "other_skills": other_skills,
        "availability": availability,
        "application_link": app_link,
        "performance_ug": perf_ug,
        "performance_pg": perf_pg,
        "performance_12": perf_12,
        "career_objective": why[:800] if why else None,
        "has_work_experience": "No",  # internship applicants often fresher; leave soft
    }


async def ensure_role(role_id: str, role_name: str, sort_order: int) -> None:
    async with async_session() as db:
        r = await db.get(HiringRole, role_id)
        if not r:
            db.add(
                HiringRole(
                    id=role_id,
                    name=role_name,
                    description="Internshala / intern list import",
                    is_active=True,
                    sort_order=sort_order,
                )
            )
            await db.commit()
            print(f"Created role {role_id}")
        else:
            print(f"Role exists: {role_id}")


async def run(reset: bool) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    roles = {
        "ai_agent_development": ("AI Agent Development", 1),
        "data_entry": ("Data Entry Internship", 40),
        "web_development_internship": ("Web Development Internship", 41),
        "telecaller": ("Telecaller / Outbound", 42),
        "internship_general": ("Internship (General)", 43),
    }
    for rid, (rname, so) in roles.items():
        await ensure_role(rid, rname, so)

    if reset:
        async with async_session() as db:
            old = (
                await db.execute(select(Candidate).where(Candidate.id.like("is_%")))
            ).scalars().all()
            for c in old:
                await db.delete(c)
            await db.commit()
            print(f"Cleared {len(old)} previous is_* Internshala excel imports")

    by_id: dict[str, dict] = {}
    for path in FILES:
        if not path.exists():
            print(f"MISSING {path.name}")
            continue
        shortlisted = "shortlisted" in path.name.lower()
        try:
            sheets = load_sheets(path)
        except Exception as e:
            print(f"ERROR {path.name}: {e}")
            continue
        n = 0
        for sn, rows in sheets:
            for row in rows:
                payload = row_to_payload(row, f"{path.name}#{sn}", shortlisted)
                if not payload:
                    continue
                cid = payload["id"]
                if cid in by_id:
                    prev = by_id[cid]
                    for k, v in payload.items():
                        if k == "id":
                            continue
                        if k == "notes" and v and v not in (prev.get("notes") or ""):
                            prev["notes"] = ((prev.get("notes") or "") + "\n" + v).strip()
                        elif k == "status" and payload["status"] == "shortlisted":
                            prev["status"] = "shortlisted"
                        elif v and not prev.get(k):
                            prev[k] = v
                else:
                    by_id[cid] = payload
                n += 1
        print(f"{path.name[:70]}: rows={n} unique_so_far={len(by_id)}")

    created = updated = 0
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    async with async_session() as db:
        for payload in by_id.values():
            cid = payload["id"]
            existing = await db.get(Candidate, cid)
            fields = {
                "role_id": payload["role_id"],
                "role_name": payload["role_name"],
                "status": payload.get("status") or "new",
                "tags": payload.get("tags") or [],
                "notes": payload.get("notes") or "",
                "name": payload.get("name"),
                "phone": payload.get("phone"),
                "email": payload.get("email"),
                "city": payload.get("city"),
                "institute": payload.get("institute"),
                "degree": payload.get("degree"),
                "stream": payload.get("stream"),
                "graduation_year": payload.get("graduation_year"),
                "other_skills": payload.get("other_skills"),
                "availability": payload.get("availability"),
                "application_link": payload.get("application_link"),
                "performance_ug": payload.get("performance_ug"),
                "performance_pg": payload.get("performance_pg"),
                "performance_12": payload.get("performance_12"),
                "career_objective": payload.get("career_objective"),
                "has_work_experience": payload.get("has_work_experience"),
                "updated_at": now,
            }
            if existing:
                for k, v in fields.items():
                    if v is not None and v != "":
                        setattr(existing, k, v)
                updated += 1
            else:
                db.add(Candidate(id=cid, created_at=now, starred=False, **fields))
                created += 1
        await db.commit()

    async with async_session() as db:
        print("\n=== Internshala excel import done ===")
        print(f"unique: {len(by_id)}  created: {created}  updated: {updated}")
        for rid in [
            "ai_agent_development",
            "data_entry",
            "web_development_internship",
            "telecaller",
            "internship_general",
        ]:
            is_cnt = (
                await db.execute(
                    select(func.count()).select_from(Candidate).where(Candidate.id.like("is_%"), Candidate.role_id == rid)
                )
            ).scalar()
            total = (
                await db.execute(
                    select(func.count()).select_from(Candidate).where(Candidate.role_id == rid)
                )
            ).scalar()
            print(f"  {rid}: is_import={is_cnt}  role_total={total}")


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--reset", action="store_true")
    args = p.parse_args()
    asyncio.run(run(args.reset))


if __name__ == "__main__":
    main()
